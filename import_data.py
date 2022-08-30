from openpyxl import load_workbook
import numpy as np
def read_date(N):
	# 打开 Excel表格
	workbook = load_workbook(filename="212.xlsx")
	# 通过 sheet 名称获取表格
	sheet = workbook["Sheet1"]
	# 获取表格的尺寸大小
	maxrow=sheet.max_row
	x=np.zeros(N,dtype=float)
	# 将数据存储在x数组中
	for i in range(N):
		j=i+1
		x[i]=sheet["A"+str(j)].value
	return x
